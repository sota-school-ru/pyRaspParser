from IPython.display import display
import pandas as pd
import openpyxl as xl
path = r'source\\test.xlsx'
wb = xl.load_workbook(path, data_only=True)
n_sheets = len(wb.worksheets)
print(n_sheets)

# !!! After using wb.worksheets return kast sheet
# sheet1 = wb.active
sheet1 = wb.worksheets[0]
print(sheet1)
sheet2 = wb.worksheets[1]
print(sheet2)

max_col = sheet1.max_column
print(max_col)
max_row = sheet1.max_row
print(max_row)

for i in range(1, max_row+1):
    for j in range(1, max_col+1):
        cell = sheet1.cell(row=i, column=j )
        print(cell.value, end=' ')
    print()

# read xlsx (not xls!!!) file into Pandas.Dataframe
print('-------info--------------')
df = pd.read_excel(path, engine='openpyxl')
display(df.info)
print('--------display(df)-------------')
display(df)
print('--------info()-------------')
print(df.info())
print('-------describe()--------------')
print(df.describe())
print('-------shape--------------')
print(df.shape)
print('-------columns()--------------')
print(df.columns)

index = df.index
print(index)

print(df['A'])

sA = df['A']
print(type(sA))

val = sA[2]
print(type(val))
print(val)

# get cell value via index
print(df.iloc [2] ['B'])
print(df.at[1, 'C'])

# get a specific row
display(df.iloc[2])

# Pandas Get the Number of Rows
display(len(df.index)) # 1
display(df.shape[0])   # 2
display(len(df.axes[0])) # 3

# Pandas Get the Number of Columns
display(df.shape[1])  # 1
display(len(df.axes[1])) # 2
