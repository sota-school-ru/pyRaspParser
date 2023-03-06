import openpyxl
import telebot

# Введите токен своего бота в Telegram
TOKEN = '6256826324:AAGJw3gRNACn1YdZ8FvPV_S3zSmOWsMjyG4'

# Создайте объект бота
bot = telebot.TeleBot(TOKEN)

# Откройте файл XLSX
wb = openpyxl.load_workbook('расписание.xlsx')

# Выберите нужный лист
ws = wb['Лист1']

# Прочитайте данные из ячеек
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    # В данном примере мы считываем значения из 2-х столбцов
    value1, value2 = row
    data.append((value1, value2))

# Закройте файл XLSX
wb.close()

# Сохраните данные в текстовый файл
with open('расписание.txt', 'w') as f:
    for row in data:
        f.write(f'{row[0]}, {row[1]}\n')

# Отправьте файл в Telegram бота
with open('расписание.txt', 'rb') as f:
    @bot.message_handler(commands=['start'])
    def start_message(message):
        bot.send_document(message.chat.id, document='расписание.txt')

    bot.infinity_polling()
    
#-------- second text -----------------------------
import openpyxl
import telebot
from telebot import types
import config

bot = telebot.TeleBot('6256826324:AAGJw3gRNACn1YdZ8FvPV_S3zSmOWsMjyG4')


@bot.message_handler(commands=['start'])
def start(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton("понедельник")
    btn2 = types.KeyboardButton("вторник")
    btn3 = types.KeyboardButton("среда")
    btn4 = types.KeyboardButton("четверг")
    btn5 = types.KeyboardButton("пятница")
    btn6 = types.KeyboardButton("суббота")
    markup.add(btn1, btn2, btn3, btn4, btn5, btn6)
    bot.send_message(message.chat.id,text="Привет, {0.first_name}! Я тестовый бот для школы СОТА".format(message.from_user),reply_markup=markup)
 

# ------------ 3-rd text -------------------------------------------------
@bot.message_handler(content_types=['text'])
def func(message):
    if message.text == "понедельник":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton("pon1a")
        btn2 = types.KeyboardButton("pon1б")
        btn3 = types.KeyboardButton("pon2а")
        btn4 = types.KeyboardButton("pon3б")
        btn5 = types.KeyboardButton("pon4а")
        btn6 = types.KeyboardButton("pon5а")
        btn7 = types.KeyboardButton("pon6а")
        btn8 = types.KeyboardButton("pon7а")
        btn9 = types.KeyboardButton("pon8а")
        btn10 = types.KeyboardButton("pon9а")
        btn11 = types.KeyboardButton("pon10а")
        back = types.KeyboardButton("Вернуться в главное меню")
        markup.add(btn1, btn2, btn3, btn4, btn5, btn6, btn7, btn8, btn9, btn10, btn11, back)
        bot.send_message(message.chat.id, text="Выберите класс!", reply_markup=markup)
        if (message.text == "pon1a"):
            button1 = types.KeyboardButton("pon 1a ur 1")
            button2 = types.KeyboardButton("pon 1a ur 2")
            button3 = types.KeyboardButton("pon 1a ur 3")
            button4 = types.KeyboardButton("pon 1a ur 4")
            button5 = types.KeyboardButton("pon 1a ur 2")
            button6 = types.KeyboardButton("суббота")
            markup.add(button1, button2, button3, button4, button5, button6)
            bot.send_message(message.chat.id, text="Вы вернулись в главное меню", reply_markup=markup)
    if message.text == "вторник":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton("vto1a")
        btn2 = types.KeyboardButton("vto1б")
        btn3 = types.KeyboardButton("vto2а")
        btn4 = types.KeyboardButton("vto3б")
        btn5 = types.KeyboardButton("vto4а")
        btn6 = types.KeyboardButton("vto5а")
        btn7 = types.KeyboardButton("vto6а")
        btn8 = types.KeyboardButton("vto7а")
        btn9 = types.KeyboardButton("vto8а")
        btn10 = types.KeyboardButton("vto9а")
        btn11 = types.KeyboardButton("vto10а")
        back = types.KeyboardButton("Вернуться в главное меню")
        markup.add(btn1, btn2, btn3, btn4, btn5, btn6, btn7, btn8, btn9, btn10, btn11, back)
        bot.send_message(message.chat.id, text="Выберите класс!", reply_markup=markup)
    elif message.text == "среда":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton("sre1a")
        btn2 = types.KeyboardButton("sre1б")
        btn3 = types.KeyboardButton("sre2а")
        btn4 = types.KeyboardButton("sre3б")
        btn5 = types.KeyboardButton("sre4а")
        btn6 = types.KeyboardButton("sre5а")
        btn7 = types.KeyboardButton("sre6а")
        btn8 = types.KeyboardButton("sre7а")
        btn9 = types.KeyboardButton("sre8а")
        btn10 = types.KeyboardButton("sre9а")
        btn11 = types.KeyboardButton("sre10а")
        back = types.KeyboardButton("Вернуться в главное меню")
        markup.add(btn1, btn2, btn3, btn4, btn5, btn6, btn7, btn8, btn9, btn10, btn11, back)
        bot.send_message(message.chat.id, text="Выберите класс!", reply_markup=markup)
    elif message.text == "четверг":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton("chet1a")
        btn2 = types.KeyboardButton("chet1б")
        btn3 = types.KeyboardButton("chet2а")
        btn4 = types.KeyboardButton("chet3б")
        btn5 = types.KeyboardButton("chet4а")
        btn6 = types.KeyboardButton("chet5а")
        btn7 = types.KeyboardButton("chet6а")
        btn8 = types.KeyboardButton("chet7а")
        btn9 = types.KeyboardButton("chet8а")
        btn10 = types.KeyboardButton("chet9а")
        btn11 = types.KeyboardButton("chet10а")
        back = types.KeyboardButton("Вернуться в главное меню")
        markup.add(btn1, btn2, btn3, btn4, btn5, btn6, btn7, btn8, btn9, btn10, btn11, back)
        bot.send_message(message.chat.id, text="Выберите класс!", reply_markup=markup)
    elif message.text == "пятница":
      markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton("pia1a")
        btn2 = types.KeyboardButton("pia1б")
        btn3 = types.KeyboardButton("pia2а")
        btn4 = types.KeyboardButton("pia3б")
        btn5 = types.KeyboardButton("pia4а")
        btn6 = types.KeyboardButton("pia5а")
        btn7 = types.KeyboardButton("pia6а")
        btn8 = types.KeyboardButton("pia7а")
        btn9 = types.KeyboardButton("pia8а")
        btn10 = types.KeyboardButton("pia9а")
        btn11 = types.KeyboardButton("pia10а")
        back = types.KeyboardButton("Вернуться в главное меню")
        markup.add(btn1, btn2, btn3, btn4, btn5, btn6, btn7, btn8, btn9, btn10, btn11, back)
        bot.send_message(message.chat.id, text="Выберите класс!", reply_markup=markup)
    elif message.text == "суббота":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton("sub1a")
        btn2 = types.KeyboardButton("sub1б")
        btn3 = types.KeyboardButton("sub2а")
        btn4 = types.KeyboardButton("sub3а")
        btn5 = types.KeyboardButton("sub4а")
        btn6 = types.KeyboardButton("sub5а")
        btn7 = types.KeyboardButton("sub6а")
        btn8 = types.KeyboardButton("sub7а")
        btn9 = types.KeyboardButton("sub8а")
        btn10 = types.KeyboardButton("sub9а")
        btn11 = types.KeyboardButton("sub10а")
        back = types.KeyboardButton("Вернуться в главное меню")
        markup.add(btn1, btn2, btn3, btn4, btn5, btn6, btn7, btn8, btn9, btn10, btn11, back)
        bot.send_message(message.chat.id, text="Выберите класс!", reply_markup=markup)
    elif (message.text == "Вернуться в главное меню"):
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        button1 = types.KeyboardButton("понедельник")
        button2 = types.KeyboardButton("вторник")
        button3 = types.KeyboardButton("среда")
        button4 = types.KeyboardButton("четверг")
        button5 = types.KeyboardButton("пятница")
        button6 = types.KeyboardButton("суббота")
        markup.add(button1, button2, button3, button4, button5, button6)
        bot.send_message(message.chat.id, text="Вы вернулись в главное меню", reply_markup=markup)



path = 'raspisanie.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active
wb.save('raspisanie.xlsx')

bot.polling(none_stop=True)
